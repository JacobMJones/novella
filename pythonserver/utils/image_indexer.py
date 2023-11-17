import os
from datetime import datetime
from PIL import Image
from openpyxl import load_workbook, Workbook
import time

folder_path = '../images' 
file_names = os.listdir(folder_path)

excel_file = '../image_data.xlsx'


# Check if the Excel file already exists
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    # Added 'timestamp' and 'id' to the header row
    ws.append(['id', 'file', 'width', 'height', 'size', 'format', 'type', 'subtype', 'tags', 'votes', 'active', 'timestamp'])

# Read existing file names from the Excel sheet to avoid duplicates
seen_files = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    seen_files.add(row[1])  # Assuming file names are in the second column

# Calculate next_id for new entries

numeric_values = [row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if isinstance(row[0], (int, float))]

if numeric_values:
    next_id = max(numeric_values) + 1
else:
    next_id = 1  # Starting ID if no numeric values are present

# Append data for new images with a timestamp
for file_name in file_names:
    if file_name not in seen_files and file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
        file_path = os.path.join(folder_path, file_name)
        try:
            with Image.open(file_path) as img:
                width, height = img.size
                file_size = os.path.getsize(file_path)
                image_format = img.format
                # Get the current timestamp
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                # Append all data including the id and timestamp
                ws.append([next_id, file_name, width, height, file_size, image_format, 'image', '', '', 0, 1, timestamp])
                next_id += 1
                seen_files.add(file_name)
                print(f'Indexed file: {file_name}')
        except IOError:
            print(f'Cannot open file: {file_name}')

excel_file_abs_path = os.path.abspath('..\image_data.xlsx')

# Get the parent directory of the Excel file
parent_dir = os.path.dirname(excel_file_abs_path)

# Construct the new file path in the parent directory
new_file_path = os.path.join(parent_dir, os.path.basename(excel_file_abs_path))

# Save the workbook to the new path
wb.save(new_file_path)
